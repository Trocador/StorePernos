# ui/ventas_view.py
from importlib.resources import path
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from ui import alerts
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class VentasView(tk.Frame):
    def __init__(self, master, controller, user):
        super().__init__(master)
        self.controller = controller
        self.user = user
        self.detalle_actual = []  # lista temporal de items antes de guardar
        self._build()

    def _build(self):
        # --- Formulario de registro ---
        # Campo para mostrar producto seleccionado
        self.producto_var = tk.StringVar(value="(ningún producto seleccionado)")
        tk.Label(self, text="Producto seleccionado:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        tk.Label(self, textvariable=self.producto_var, anchor="w", width=50, relief="sunken").grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        tk.Button(self, text="Buscar producto", command=self._abrir_buscador).grid(row=0, column=2, padx=5)

        self.precio_unidad_var = tk.DoubleVar(value=0.0)
        self.precio_kilo_var = tk.DoubleVar(value=0.0)

        tk.Label(self, text="Precio por unidad").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        tk.Label(self, textvariable=self.precio_unidad_var, relief="sunken", width=20).grid(row=1, column=1, sticky="w", padx=5)

        tk.Label(self, text="Precio por kilo").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        tk.Label(self, textvariable=self.precio_kilo_var, relief="sunken", width=20).grid(row=2, column=1, sticky="w", padx=5)
        
        frame_cantidad_total = tk.Frame(self)
        frame_cantidad_total.grid(row=1, column=0, columnspan=3, pady=5)

        tk.Label(frame_cantidad_total, text="Cantidad").pack(side="left", padx=5)
        self.cantidad_var = tk.IntVar(value=1)
        tk.Spinbox(frame_cantidad_total, from_=1, to=1000, textvariable=self.cantidad_var, width=10).pack(side="left", padx=5)

        tk.Label(frame_cantidad_total, text="Total a pagar").pack(side="left", padx=5)
        self.total_var = tk.DoubleVar(value=0.0)
        tk.Entry(frame_cantidad_total, textvariable=self.total_var, state="readonly", width=15).pack(side="left", padx=5)

        tk.Label(self, text="Tipo venta").grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.tipo_venta_var = tk.StringVar(value="unidad")
        ttk.Combobox(self, textvariable=self.tipo_venta_var, values=["unidad","kilo"], state="readonly").grid(row=3, column=1)

        tk.Button(self, text="Agregar ítem", command=self._agregar_item).grid(row=4, column=0, pady=10)
        tk.Button(self, text="Resetear ítems", command=self._reset_items).grid(row=4, column=2, pady=10)
        tk.Button(self, text="Finalizar venta", command=self._registrar).grid(row=4, column=1, pady=10)
        # --- Listado de ventas ---
        self.tree_ventas = ttk.Treeview(self, columns=("id", "usuario", "fecha", "total"), show="headings")
        self.tree_ventas.heading("id", text="ID Venta")
        self.tree_ventas.heading("usuario", text="Usuario")
        self.tree_ventas.heading("fecha", text="Fecha")
        self.tree_ventas.heading("total", text="Total")
        self.tree_ventas.grid(row=5, column=0, columnspan=2, sticky="nsew", pady=10)

        scrollbar1 = ttk.Scrollbar(self, orient="vertical", command=self.tree_ventas.yview)
        self.tree_ventas.configure(yscroll=scrollbar1.set)
        scrollbar1.grid(row=5, column=2, sticky="ns")

        # --- Listado de detalle ---
        self.tree_detalle = ttk.Treeview(
            self,
            columns=("id_detalle", "id_producto", "producto", "cantidad", "tipo_venta", "precio_unitario", "subtotal"),
            show="headings"
        )
        self.tree_detalle.heading("id_detalle", text="ID Detalle")
        self.tree_detalle.heading("id_producto", text="ID Producto")
        self.tree_detalle.heading("producto", text="Producto")
        self.tree_detalle.heading("cantidad", text="Cantidad")
        self.tree_detalle.heading("tipo_venta", text="Tipo Venta")
        self.tree_detalle.heading("precio_unitario", text="Precio Unitario")
        self.tree_detalle.heading("subtotal", text="Subtotal")
        self.tree_detalle.grid(row=6, column=0, columnspan=2, sticky="nsew", pady=10)

        scrollbar2 = ttk.Scrollbar(self, orient="vertical", command=self.tree_detalle.yview)
        self.tree_detalle.configure(yscroll=scrollbar2.set)
        scrollbar2.grid(row=6, column=2, sticky="ns")
        
        # 🔒 PANEL SOLO ADMIN
        if self.user["rol"] == "admin":
            frame_reportes = tk.LabelFrame(self, text="Reportes de Control (Admin)")
            frame_reportes.grid(row=7, column=0, columnspan=3, sticky="ew", pady=10)

            tk.Button(frame_reportes, text="Reporte del Día", command=self._reporte_dia).pack(side="left", padx=10)
            tk.Button(frame_reportes, text="Reporte Semanal", command=self._reporte_semana).pack(side="left", padx=10)


        # Configuración de expansión
        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(6, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # Evento de selección
        self.tree_ventas.bind("<<TreeviewSelect>>", self._mostrar_detalle)

        # Cargar ventas iniciales
        self._cargar_ventas()

    def _agregar_item(self):
        producto_texto = self.producto_var.get()
        if not producto_texto or producto_texto == "(ningún producto seleccionado)":
            alerts.error("Debe seleccionar un producto")
            return
        id_producto = int(producto_texto.split(" - ")[0])

        if self.tipo_venta_var.get() == "unidad":
            precio = self.precio_unidad_var.get()
        else:
            precio = self.precio_kilo_var.get()

        item = {
            "id_producto": id_producto,
            "cantidad": self.cantidad_var.get(),
            "tipo_venta": self.tipo_venta_var.get(),
            "precio_unitario": precio,
            "subtotal": self.cantidad_var.get() * precio
        }

        self.detalle_actual.append(item)
        self.total_var.set(sum(i["subtotal"] for i in self.detalle_actual))
        alerts.info(f"Item agregado: {item}")

    def _registrar(self):
        if not self.detalle_actual:
            alerts.error("Debe agregar al menos un ítem")
            return

        ok = self.controller.registrar(self.user["id_usuario"], self.detalle_actual)
        if ok:
            alerts.info("Venta registrada correctamente")
            self.detalle_actual.clear()
            self.total_var.set(0.0)
            self._cargar_ventas()
        else:
            alerts.error("Error al registrar venta")

    def _cargar_ventas(self):
        for row in self.tree_ventas.get_children():
            self.tree_ventas.delete(row)

        ventas = self.controller.listar()
        for v in ventas:
            self.tree_ventas.insert("", "end", values=(v["id_venta"], v["id_usuario"], v["fecha"], v["total"]))

    def _mostrar_detalle(self, event):
        for row in self.tree_detalle.get_children():
            self.tree_detalle.delete(row)

        selected = self.tree_ventas.selection()
        if not selected:
            return
        id_venta = self.tree_ventas.item(selected[0])["values"][0]

        detalles = self.controller.detalle(id_venta)
        for d in detalles:
            self.tree_detalle.insert("", "end", values=(
                d["id_detalle"],
                d["id_producto"],
                d["producto"],   # 🔥 ahora muestra nombre
                d["cantidad"],
                d["tipo_venta"],
                d["precio_unitario"],
                d["subtotal"]
            ))

    def _reset_items(self):
        self.detalle_actual.clear()
        self.total_var.set(0.0)
        alerts.info("Ítems de la venta reseteados")

    def _exportar_excel(self, resumen, ventas):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Ventas"

        # Encabezado resumen
        ws.append(["Tipo Reporte", resumen["tipo"]])
        ws.append(["Desde", resumen["fecha_inicio"]])
        ws.append(["Hasta", resumen["fecha_fin"]])
        ws.append([])
        ws.append(["Total Vendido", resumen["total_vendido"]])
        ws.append(["Ganancia (50%)", resumen["ganancia"]])
        ws.append([])

        # Detalle de ventas
        for v in ventas:
            venta = v["venta"]
            ws.append([f"Venta ID {venta['id_venta']} - Usuario {venta['id_usuario']} - Fecha {venta['fecha']} - Total {venta['total']}"])
            ws.append(["ID Detalle", "ID Producto", "Producto", "Cantidad", "Tipo Venta", "Precio Unitario", "Subtotal"])
            for d in v["detalles"]:
                ws.append([
                    d["id_detalle"],
                    d["id_producto"],
                    d["producto"],
                    d["cantidad"],
                    d["tipo_venta"],
                    d["precio_unitario"],
                    d["subtotal"]
                ])
            ws.append([])  # línea en blanco entre ventas

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(path)
        alerts.info("Reporte exportado correctamente")

    def _reporte_dia(self):
        resumen = self.controller.resumen_dia()
        ventas = self.controller.ventas_dia()
        ventas_detalladas = self.controller.ventas_con_detalle(ventas)
        alerts.info(f"Total vendido hoy: {resumen['total_vendido']} | Ganancia: {resumen['ganancia']}")
        self._exportar_excel(resumen, ventas_detalladas)

    def _reporte_semana(self):
        resumen = self.controller.resumen_semana()
        ventas = self.controller.ventas_semana()
        ventas_detalladas = self.controller.ventas_con_detalle(ventas)
        alerts.info(f"Total vendido semana: {resumen['total_vendido']} | Ganancia: {resumen['ganancia']}")
        self._exportar_excel(resumen, ventas_detalladas)

    def _abrir_buscador(self):
        popup = tk.Toplevel(self)
        popup.title("Buscar producto")
        popup.geometry("700x400")

        tk.Label(popup, text="Filtrar:").pack(pady=5)
        filtro_var = tk.StringVar()
        filtro_entry = tk.Entry(popup, textvariable=filtro_var)
        filtro_entry.pack(fill="x", padx=10)

        tree = ttk.Treeview(popup, columns=("id", "nombre", "stock"), show="headings")
        tree.heading("id", text="ID")
        tree.heading("nombre", text="Producto")
        tree.heading("stock", text="Stock")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        productos = self.controller.listar_productos()

        def cargar_productos(filtro=""):
            for row in tree.get_children():
                tree.delete(row)
            for p in productos:
                nombre = f"{p['tipo']} {p['medida']} {p['largo'] or ''}"
                if filtro.lower() in nombre.lower():
                    tree.insert("", "end", values=(p["id_producto"], nombre, p["stock"]))

        cargar_productos()
        filtro_var.trace_add("write", lambda *args: cargar_productos(filtro_var.get()))

        def seleccionar():
            sel = tree.selection()
            if sel:
                values = tree.item(sel[0])["values"]
                id_producto = values[0]
                producto = self.controller.get_producto(id_producto)
                # 🔥 Mostrar producto seleccionado
                self.producto_var.set(f"{id_producto} - {values[1]}")
                # 🔥 Mostrar precios automáticamente
                self.precio_unidad_var.set(producto["precio_unidad"])
                self.precio_kilo_var.set(producto["precio_kilo"])
                popup.destroy()

        tk.Button(popup, text="Seleccionar", command=seleccionar).pack(pady=5)